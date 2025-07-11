from __future__ import annotations

import time
import traceback
from datetime import datetime
from os import path, remove
from shutil import copy2
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from PyQt6.QtCore import QObject, QReadWriteLock

from .logger import get_logger
from .path_utils import make_relative_path, normalize_path, split_drive_or_unc
from .performance_profiler import PerformanceProfiler, profile_method


class ExcelManager(QObject):
    """Manages Excel file operations and data caching."""

    def __init__(self) -> None:
        super().__init__()
        self.excel_data: Optional[pd.DataFrame] = None
        # Cache format: {(row_idx: int, col_idx: int): hyperlink_target: Optional[str]}
        self._hyperlink_cache: Dict[Tuple[int, int], Optional[str]] = {}
        self._cache_lock = QReadWriteLock()
        self._last_file: Optional[str] = None
        self._last_sheet: Optional[str] = None
        self._sheet_cache: Dict[str, list] = {}  # Cache for sheet names
        self._column_cache: Dict[str, list] = {}  # Cache for column names
        self._header_cache: Dict[str, Dict[str, int]] = {} # Cache for header {sheet_key: {col_name: col_idx}}

        # Performance profiler
        self._profiler = PerformanceProfiler()

        # Workbook caching for performance
        self._cached_workbook = None
        self._cached_workbook_path = None

    def clear_caches(self) -> None:
        """Clear all cached data."""
        logger = get_logger()
        self._cache_lock.lockForWrite()
        try:
            self._hyperlink_cache.clear()
            self._last_file = None
            self._last_sheet = None
            self._sheet_cache.clear()
            self._column_cache.clear()
            self._header_cache.clear()
            self.excel_data = None
            # Clear workbook cache
            if self._cached_workbook:
                try:
                    self._cached_workbook.close()
                except:
                    pass
                self._cached_workbook = None
                self._cached_workbook_path = None
            logger.debug("All Excel caches cleared")
        finally:
            self._cache_lock.unlock()

    def _get_cached_workbook(self, file_path: str, use_cache: bool = True):
        """Get a workbook, using cache if available and requested."""
        if use_cache and self._cached_workbook and self._cached_workbook_path == file_path:
            logger = get_logger()
            logger.debug(f"Using cached workbook for {file_path}")
            return self._cached_workbook, False  # False = not newly loaded

        # Load new workbook
        if self._cached_workbook and self._cached_workbook_path != file_path:
            # Close previous workbook
            try:
                self._cached_workbook.close()
            except Exception:
                pass

        self._profiler.start_operation("load_workbook")
        wb = load_workbook(file_path)
        self._profiler.end_operation("load_workbook")

        if use_cache:
            self._cached_workbook = wb
            self._cached_workbook_path = file_path

        return wb, True  # True = newly loaded

    def clear_workbook_cache(self) -> None:
        """Clear the workbook cache to free memory."""
        if self._cached_workbook:
            try:
                self._cached_workbook.close()
            except Exception:
                pass
            self._cached_workbook = None
            self._cached_workbook_path = None
            logger = get_logger()
            logger.debug("Workbook cache cleared")

    def load_excel_data(
        self, file_path: str, sheet_name: str, force_reload: bool = False
    ) -> bool:
        """Load Excel data from file into DataFrame.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to load
            force_reload: If True, will reload data even if it's already cached

        Returns:
            bool: True if data was reloaded, False if using cached data
        """
        if not file_path or not sheet_name:
            return False

        try:
            logger = get_logger()
            self._profiler.start_operation("load_excel_data")

            # Check if we need to reload
            if (
                not force_reload
                and self.excel_data is not None
                and self._last_file == file_path
                and self._last_sheet == sheet_name
            ):
                logger.debug(f"Using cached Excel data (force_reload={force_reload})")
                self._profiler.end_operation("load_excel_data")
                return False

            logger.debug(f"Loading Excel data from {file_path}, sheet: {sheet_name}")

            # Try to normalize path for network paths
            normalized_path = file_path
            # Handle Windows UNC paths
            if file_path.startswith("//") or file_path.startswith("\\\\"):
                try:
                    # Make sure the path is in a consistent format
                    parts = file_path.replace("/", "\\").strip("\\").split("\\")
                    if len(parts) >= 2:
                        normalized_path = f"\\\\{parts[0]}\\{parts[1]}"
                        if len(parts) > 2:
                            normalized_path += "\\" + "\\".join(parts[2:])
                    logger.debug(f"Normalized network path: {normalized_path}")
                except Exception as path_err:
                    logger.warning(f"Path normalization error: {str(path_err)}")

            # Load the data
            try:
                self.excel_data = pd.read_excel(
                    normalized_path, sheet_name=sheet_name, engine="openpyxl"
                )
            except OSError:
                logger.debug("Failed to read with normalized path, trying original path")
                # If normalized path fails, try the original path
                self.excel_data = pd.read_excel(
                    file_path, sheet_name=sheet_name, engine="openpyxl"
                )

            # Update last loaded file info
            self._last_file = file_path
            self._last_sheet = sheet_name

            # Clear hyperlink cache (acquire lock)
            self._cache_lock.lockForWrite()
            try:
                self._hyperlink_cache.clear()
                logger.debug("Hyperlink cache cleared during data load")
            finally:
                self._cache_lock.unlock()

            # Clear header cache for this sheet
            sheet_key = f"{file_path}:{sheet_name}"
            if sheet_key in self._header_cache:
                del self._header_cache[sheet_key]

            self._profiler.end_operation("load_excel_data")
            return True

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error loading Excel data: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")

            self.excel_data = None
            self._last_file = None
            self._last_sheet = None
            self._cache_lock.lockForWrite()
            try:
                self._hyperlink_cache.clear()
            finally:
                self._cache_lock.unlock()
            self._profiler.end_operation("load_excel_data")
            raise

    def preload_hyperlinks_async(
        self,
        file_path: str,
        sheet_name: str,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> None:
        """Preload all hyperlinks from the sheet into the cache asynchronously."""
        logger = get_logger()
        logger.info(f"Starting hyperlink preloading for {file_path} - {sheet_name}")
        self._cache_lock.lockForWrite()
        try:
            self._hyperlink_cache.clear()

            # Try to normalize path for network paths
            normalized_path = file_path
            if file_path.startswith("//") or file_path.startswith("\\\\"):
                try:
                    parts = file_path.replace("/", "\\").strip("\\").split("\\")
                    if len(parts) >= 2:
                        normalized_path = f"\\\\{parts[0]}\\{parts[1]}"
                        if len(parts) > 2:
                            normalized_path += "\\" + "\\".join(parts[2:])
                    logger.debug(f"Normalized network path for preloading: {normalized_path}")
                except Exception as path_err:
                    logger.warning(f"Path normalization error during preloading: {str(path_err)}")

            try:
                # Load workbook WITHOUT read_only=True to access hyperlinks
                wb = load_workbook(normalized_path, data_only=True)
            except Exception:
                logger.debug(f"Failed preload with normalized path, trying original: {file_path}")
                try:
                    # Load workbook WITHOUT read_only=True to access hyperlinks
                    wb = load_workbook(file_path, data_only=True)
                except Exception as wb_err:
                    logger.error(f"Could not load workbook for preloading: {str(wb_err)}")
                    if hasattr(wb_err, "__traceback__"):
                        logger.error(f"Traceback: {traceback.format_exception(type(wb_err), wb_err, wb_err.__traceback__)}")
                    return # Exit if workbook can't be loaded

            try:
                ws = wb[sheet_name]
            except KeyError:
                logger.warning(f"Sheet '{sheet_name}' not found during preloading.")
                wb.close()
                return
            except Exception as sheet_err:
                logger.error(f"Error accessing sheet during preloading: {str(sheet_err)}")
                if hasattr(sheet_err, "__traceback__"):
                    logger.error(f"Traceback: {traceback.format_exception(type(sheet_err), sheet_err, sheet_err.__traceback__)}")
                wb.close()
                return

            total_rows = ws.max_row - 1 # Exclude header row
            processed_rows = 0

            # Iterate through all cells to find hyperlinks
            for row_idx, row in enumerate(ws.iter_rows(min_row=2)): # 0-based index from iter_rows
                for col_idx, cell in enumerate(row): # 0-based index from row iteration
                    if cell.hyperlink:
                        try:
                            target = cell.hyperlink.target
                            # Store with 0-based indices
                            cache_key = (row_idx, col_idx)
                            self._hyperlink_cache[cache_key] = target
                            # --- Add Debug Print ---
                            # logger.debug(f"Added to cache: key={cache_key}, target='{target}'")
                            # --- End Debug Print ---
                        except Exception as link_err:
                            logger.warning(f"Error reading hyperlink at ({row_idx}, {col_idx}): {str(link_err)}")
                            self._hyperlink_cache[(row_idx, col_idx)] = None # Mark as checked but failed

                processed_rows += 1
                if progress_callback and total_rows > 0 and processed_rows % 100 == 0: # Report every 100 rows
                    progress = int((processed_rows / total_rows) * 100)
                    progress_callback(progress)

            wb.close()
            logger.info(f"Finished hyperlink preloading. Cached {len(self._hyperlink_cache)} links.")
            if progress_callback:
                progress_callback(100) # Signal completion

        except Exception as e:
            logger.error(f"Error during hyperlink preloading: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
            # Keep potentially partially filled cache? Or clear? Clearing seems safer.
            self._hyperlink_cache.clear()
        finally:
            self._cache_lock.unlock()

    def refresh_hyperlink_cache(
        self,
        file_path: str,
        sheet_name: str,
        progress_callback: Optional[Callable[[int], None]] = None,
    ) -> None:
        """Force a refresh of the hyperlink cache."""
        logger = get_logger()
        logger.info("Refreshing hyperlink cache...")
        # This can be called directly, potentially in a separate thread if needed
        self.preload_hyperlinks_async(file_path, sheet_name, progress_callback)

    def get_hyperlink(self, row_idx: int, col_idx: int) -> Optional[str]:
        """Get hyperlink target from cache using 0-based indices.

        Args:
            row_idx: 0-based row index.
            col_idx: 0-based column index.

        Returns:
            Optional[str]: Hyperlink target or None if not found/cached.
        """
        self._cache_lock.lockForRead()
        try:
            link = self._hyperlink_cache.get((row_idx, col_idx))
            # logger = get_logger()
            # logger.debug(f"Cache lookup for ({row_idx}, {col_idx}): {'Found' if link else 'Not Found'}")
            return link
        finally:
            self._cache_lock.unlock()

    def _get_column_index(self, file_path: str, sheet_name: str, column_name: str) -> Optional[int]:
        """Get the 0-based index for a column name, using cache."""
        sheet_key = f"{file_path}:{sheet_name}"
        if sheet_key not in self._header_cache:
            try:
                wb = load_workbook(file_path, read_only=True)
                ws = wb[sheet_name]
                header = {cell.value: idx for idx, cell in enumerate(ws[1])} # 0-based index
                self._header_cache[sheet_key] = header
                wb.close()
            except Exception as e:
                logger = get_logger()
                logger.error(f"Error caching header for {sheet_key}: {str(e)}")
                if hasattr(e, "__traceback__"):
                    logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
                return None

        return self._header_cache[sheet_key].get(column_name)

    def update_pdf_link(
        self,
        file_path: str,
        sheet_name: str,
        row_idx: int, # 0-based row index
        pdf_path: str,
        column_name: str,
    ) -> Optional[str]:
        """Update PDF hyperlink in Excel file.

        Returns:
            Optional[str]: Original hyperlink if there was one, None otherwise
        """
        try:
            logger = get_logger()
            logger.info(
                f"Updating PDF link in Excel: file={file_path}, sheet={sheet_name}, row_idx={row_idx}, column={column_name}"
            )

            # Normalize paths using path_utils
            normalized_excel_path = normalize_path(file_path)
            normalized_pdf_path = normalize_path(pdf_path)

            logger.debug(f"Linking to PDF: {normalized_pdf_path}")

            # Load workbook (with caching for performance)
            wb, newly_loaded = self._get_cached_workbook(file_path, use_cache=True)
            ws = wb[sheet_name]

            # Find the column index (0-based) using helper
            col_idx_0_based = self._get_column_index(file_path, sheet_name, column_name)
            if col_idx_0_based is None:
                raise ValueError(f"Column '{column_name}' not found")
            col_idx_1_based = col_idx_0_based + 1 # openpyxl uses 1-based index

            # Get the cell
            excel_row_1_based = row_idx + 2  # +1 for header, +1 for 1-based index
            cell = ws.cell(row=excel_row_1_based, column=col_idx_1_based)

            logger.debug(f"Excel cell: {cell.coordinate}, Current value: {cell.value}")

            # Store original hyperlink
            original_link = None
            if cell.hyperlink:
                original_link = (
                    cell.hyperlink.target
                    if hasattr(cell.hyperlink, "target")
                    else str(cell.hyperlink)
                )
                logger.debug(f"Found existing hyperlink: {original_link}")

            # Use path_utils to determine if paths are on same drive/mount
            excel_drive, _ = split_drive_or_unc(normalized_excel_path)
            pdf_drive, _ = split_drive_or_unc(normalized_pdf_path)

            logger.debug(f"Excel mount: {excel_drive}, PDF mount: {pdf_drive}")

            # Calculate target path using path_utils make_relative_path
            target_path = make_relative_path(normalized_excel_path, normalized_pdf_path)
            logger.debug(f"Target path for hyperlink: {target_path}")

            # Update hyperlink using Excel formula and styling
            hyperlink = Hyperlink(ref=cell.coordinate, target=target_path)
            hyperlink.target_mode = "file"
            cell.hyperlink = hyperlink
            cell.style = "Hyperlink"

            # Save workbook
            wb.save(file_path)
            logger.debug(f"Set HYPERLINK formula: {cell.value}")
            logger.info(f"Updated Excel hyperlink, original: {original_link}")
            logger.debug("Excel file saved successfully")

            # Update cache (acquire lock)
            self._cache_lock.lockForWrite()
            try:
                # Use 0-based indices for cache key
                self._hyperlink_cache[(row_idx, col_idx_0_based)] = target_path
                logger.debug(f"Updated cache for ({row_idx}, {col_idx_0_based})")
            finally:
                self._cache_lock.unlock()

            return original_link

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error updating PDF link: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
            return None

    def revert_pdf_link(
        self,
        excel_file: str,
        sheet_name: str,
        row_idx: int, # 0-based row index
        filter2_col: str,
        original_hyperlink: Optional[str],
        original_value: str,
    ) -> bool:
        """Revert PDF hyperlink in Excel file.

        Args:
            excel_file: Path to the Excel file.
            sheet_name: Name of the sheet.
            row_idx: 0-based row index.
            filter2_col: Name of the filter2 column.
            original_hyperlink: Original hyperlink to restore, or None to remove hyperlink.
            original_value: Original value to restore.

        Returns:
            bool: True if successful, False otherwise.
        """
        try:
            logger = get_logger()
            logger.info(
                f"Reverting hyperlink for row {row_idx} in {excel_file}, sheet {sheet_name}"
            )
            logger.debug(
                f"Original hyperlink: {original_hyperlink}, Original value: {original_value}"
            )

            wb = load_workbook(excel_file)
            ws = wb[sheet_name]

            # Find the column index (0-based) using helper
            col_idx_0_based = self._get_column_index(excel_file, sheet_name, filter2_col)
            if col_idx_0_based is None:
                raise ValueError(f"Column '{filter2_col}' not found")
            col_idx_1_based = col_idx_0_based + 1 # openpyxl uses 1-based index

            # Get the cell
            excel_row_1_based = row_idx + 2 # +1 header, +1 for 1-based index
            cell = ws.cell(row=excel_row_1_based, column=col_idx_1_based)
            logger.debug(
                f"Identified Excel cell: {cell.coordinate}, Current value: {cell.value}"
            )

            # Update or remove hyperlink - handle different openpyxl versions
            if original_hyperlink:
                try:
                    # Method 1: Using Hyperlink class (newer versions of openpyxl)
                    hyperlink = Hyperlink(
                        ref=cell.coordinate, target=original_hyperlink
                    )
                    hyperlink.target_mode = "file"
                    cell.hyperlink = hyperlink
                    logger.debug(
                        f"Restored hyperlink using newer API: {original_hyperlink}"
                    )
                except TypeError as e:
                    try:
                        # Method 2: Direct assignment (older versions)
                        cell.hyperlink = original_hyperlink
                        logger.debug(
                            f"Restored hyperlink using direct assignment: {original_hyperlink}"
                        )
                    except Exception as e2:
                        logger.warning(
                            f"Hyperlink methods failed: {str(e)}, {str(e2)}. Setting display text only."
                        )
                        # Fallback: Set display text only
                        cell.value = f"{original_value} [Link:{original_hyperlink}]"
                        logger.debug("Set display text with link reference")
            else:
                # Remove hyperlink - handle different openpyxl versions
                try:
                    cell.hyperlink = None
                    logger.debug("Removed hyperlink")
                except Exception as e:
                    # If direct removal fails, try other methods
                    try:
                        # Some versions may use cell._hyperlink
                        if hasattr(cell, "_hyperlink"):
                            cell._hyperlink = None
                            logger.debug(
                                "Removed hyperlink using _hyperlink attribute"
                            )
                    except Exception as e2:
                        logger.warning(
                            f"Failed to remove hyperlink: {str(e)}, {str(e2)}"
                        )

            # Ensure cell value is restored
            if cell.value != original_value:
                cell.value = original_value
                logger.debug(f"Restored cell value to: {original_value}")

            # Save workbook
            wb.save(excel_file)
            logger.debug("Excel file saved successfully after reversion")

            # Update cache (acquire lock)
            self._cache_lock.lockForWrite()
            try:
                # Use 0-based indices for cache key
                cache_key = (row_idx, col_idx_0_based)
                if original_hyperlink:
                    self._hyperlink_cache[cache_key] = original_hyperlink
                    logger.debug(f"Updated cache for {cache_key} with original link")
                elif cache_key in self._hyperlink_cache:
                    del self._hyperlink_cache[cache_key]
                    logger.debug(f"Removed cache entry for {cache_key}")
            finally:
                self._cache_lock.unlock()

            return True

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error reverting PDF link: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
            return False

    def get_available_sheets(self, file_path: str) -> list[str]:
        """Get list of available sheets in Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        try:
            # Check cache first
            if file_path in self._sheet_cache:
                return self._sheet_cache[file_path]

            # Load workbook
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames

            # Cache results
            self._sheet_cache[file_path] = sheet_names

            return sheet_names

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error getting sheet names: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
            raise

    def get_sheet_columns(self, file_path: str, sheet_name: str) -> list[str]:
        """Get list of column names from specified sheet.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read

        Returns:
            List of column names from first row
        """
        try:
            # Generate cache key
            cache_key = f"{file_path}:{sheet_name}"

            # Check cache first
            if cache_key in self._column_cache:
                return self._column_cache[cache_key]

            # Load workbook and get first row only
            wb = load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]

            # Get header row values
            header_row = next(ws.rows)
            column_names = [cell.value for cell in header_row if cell.value]

            # Cache results
            self._column_cache[cache_key] = column_names

            return column_names

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error getting column names: {str(e)}")
            if hasattr(e, "__traceback__"):
                logger.error(f"Traceback: {traceback.format_exception(type(e), e, e.__traceback__)}")
            raise

    def add_new_row(
        self,
        file_path: str,
        sheet_name: str,
        columns: List[str],
        values: List[str],
        create_backup: bool = True,
    ) -> Tuple[Dict[str, Any], int]:
        """Add a new row to the Excel file.

        Returns:
            Tuple[Dict[str, Any], int]: Dictionary of row data and the row index
        """
        try:
            logger = get_logger()
            self._profiler.start_operation("add_new_row")
            logger.debug(
                f"Cache state before adding new row - size: {len(self._hyperlink_cache)}"
            )

            # Validate input
            if len(columns) != len(values):
                raise ValueError(
                    f"Number of columns ({len(columns)}) and values ({len(values)}) must match"
                )

            logger.info(f"Adding new row with values: {dict(zip(columns, values))}")

            # Create a backup (optional for performance)
            backup_file = None
            if create_backup:
                self._profiler.start_operation("create_backup")
                backup_file = file_path + ".bak"
                copy2(file_path, backup_file)
                logger.debug(f"Created backup at {backup_file}")
                self._profiler.end_operation("create_backup")
            else:
                logger.debug("Skipping backup creation for performance")

            try:
                # Load workbook (with caching for performance)
                wb, newly_loaded = self._get_cached_workbook(file_path, use_cache=True)
                ws = wb[sheet_name]

                # Get header row
                header_row = ws[1]
                col_indices = {
                    cell.value: idx + 1 for idx, cell in enumerate(header_row)
                }

                # Verify all filter columns exist
                for col in columns:
                    if col not in col_indices:
                        raise ValueError(
                            f"Column '{col}' not found in Excel file. Available columns: {', '.join(col_indices.keys())}"
                        )

                # Find the first table's range to determine where to add the new row
                table_end_row = None
                for table in ws.tables.values():
                    try:
                        current_ref = table.ref
                        ref_parts = current_ref.split(":")
                        if len(ref_parts) == 2:
                            end_ref = ref_parts[1]
                            table_end_row = int("".join(filter(str.isdigit, end_ref)))
                            logger.debug(f"Found table with end row: {table_end_row}")
                            break  # Use the first table found
                    except Exception as e:
                        logger.error(f"Error processing table reference: {str(e)}")
                        continue

                # If we found a table, add the row immediately after it
                if table_end_row:
                    new_row = table_end_row + 1
                else:
                    # Fallback to adding at the end if no table found
                    new_row = ws.max_row + 1

                logger.debug(
                    f"Adding row at index {new_row - 2} (Excel row {new_row})"
                )

                # First pass: Copy all formats from the template row (use row before new row)
                template_row = new_row - 1
                logger.debug(f"Using template row {template_row} for formatting")

                # Second pass: Set values with proper type conversion
                for col, val in zip(columns, values):
                    col_idx = col_indices[col]
                    # Get the cell in the new row
                    new_cell = ws.cell(row=new_row, column=col_idx)

                    # Convert value based on the column type
                    if "DATE" in col.upper() and val:
                        try:
                            # Try to parse date in various formats
                            date_formats = [
                                "%d/%m/%Y",
                                "%d-%m-%Y",
                                "%Y-%m-%d",
                                "%Y/%m/%d",
                            ]
                            date_val = None

                            for fmt in date_formats:
                                try:
                                    date_val = datetime.strptime(val, fmt)
                                    break
                                except ValueError:
                                    continue

                            if date_val:
                                new_cell.value = date_val
                                new_cell.number_format = "DD/MM/YYYY"
                            else:
                                # Fallback to pandas datetime parsing
                                date_val = pd.to_datetime(val)
                                new_cell.value = date_val.to_pydatetime()
                                new_cell.number_format = "DD/MM/YYYY"
                        except Exception as e:
                            logger.warning(
                                f"Could not parse date '{val}' for column '{col}': {str(e)}"
                            )
                            new_cell.value = val
                    elif "MNT" in col.upper() or "MONTANT" in col.upper():
                        try:
                            # Handle number format
                            num_str = val.replace(" ", "").replace(",", ".")
                            num_val = float(num_str)
                            new_cell.value = num_val
                            new_cell.number_format = '_-* #,##0.00\\ _€_-;\\-* #,##0.00\\ _€_-;_-* "-"??\\ _€_-;_-@_-'
                            new_cell.style = "Comma"
                        except Exception as e:
                            new_cell.value = val
                            logger.warning(
                                f"Could not parse number '{val}' for column '{col}': {str(e)}"
                            )
                    else:
                        new_cell.value = val

                    logger.debug(f"Set value '{val}' for column '{col}'")

                # Check and expand table ranges to include the new row
                if table_end_row:
                    logger.debug("Checking for tables that need to be expanded")
                    for table in ws.tables.values():
                        try:
                            current_ref = table.ref
                            # Split table reference into components (e.g., 'A1:D10' -> ['A1', 'D10'])
                            ref_parts = current_ref.split(":")
                            if len(ref_parts) != 2:
                                continue

                            start_ref, end_ref = ref_parts

                            # Extract row numbers from references
                            start_row = int("".join(filter(str.isdigit, start_ref)))
                            end_row = int("".join(filter(str.isdigit, end_ref)))

                            # Check if new row is immediately after table
                            if end_row == new_row - 1:
                                # Get column letters from references (e.g., 'A' from 'A1')
                                start_col = "".join(filter(str.isalpha, start_ref))
                                end_col = "".join(filter(str.isalpha, end_ref))

                                # Create new reference that includes the new row
                                new_ref = f"{start_col}{start_row}:{end_col}{new_row}"
                                table.ref = new_ref
                                logger.debug(
                                    f"Expanded table '{table.displayName}' range to {new_ref}"
                                )
                        except Exception as table_e:
                            logger.warning(f"Error expanding table: {str(table_e)}")
                            # Continue with other tables even if one fails
                            continue

                # Save workbook
                self._profiler.start_operation("save_workbook")
                wb.save(file_path)
                self._profiler.end_operation("save_workbook")

                # Update cache for the new row
                self._hyperlink_cache[new_row - 2] = False

                # Create a row data dictionary that includes all the values
                row_data = {}
                if self.excel_data is not None:
                    # Create data for all columns (initialize with None)
                    for col in self.excel_data.columns:
                        row_data[col] = None

                    # Update with the values we have
                    for col, val in zip(columns, values):
                        row_data[col] = val

                    # Add the new row to our DataFrame
                    self.excel_data = pd.concat(
                        [self.excel_data, pd.DataFrame([row_data])], ignore_index=True
                    )

                    # Row index is the last row (0-based)
                    row_idx = len(self.excel_data) - 1
                else:
                    # If DataFrame doesn't exist yet, reload from Excel
                    self.load_excel_data(file_path, sheet_name)
                    # Get the new row index (Excel row minus header and 1-based index)
                    row_idx = new_row - 2
                    # Get row data
                    if row_idx < len(self.excel_data):
                        row_data = self.excel_data.iloc[row_idx].to_dict()
                    else:
                        # Manual construction if row index is out of bounds
                        row_data = {col: val for col, val in zip(columns, values)}

                # Remove backup after successful write
                if backup_file and path.exists(backup_file):
                    remove(backup_file)
                    logger.debug("Removed backup file after successful write")

                logger.info(f"Successfully added new row at index {row_idx}")
                self._profiler.end_operation("add_new_row")
                return row_data, row_idx

            finally:
                # Only close workbook if we're not caching it
                if "wb" in locals() and wb != self._cached_workbook:
                    wb.close()

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error adding new row: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")

            # Try to restore from backup
            if "backup_file" in locals() and backup_file and path.exists(backup_file):
                try:
                    copy2(backup_file, file_path)
                    logger.info("Restored from backup after error")
                except Exception as backup_e:
                    logger.error(f"Failed to restore from backup: {str(backup_e)}")
                    if hasattr(backup_e, "__traceback__"):
                        logger.error(f"Backup restore traceback: {traceback.format_exception(type(backup_e), backup_e, backup_e.__traceback__)}")

            self._profiler.end_operation("add_new_row")
            raise

    def remove_row(
        self,
        file_path: str,
        sheet_name: str,
        row_idx: int,  # 0-based row index
    ) -> bool:
        """Remove a row from the Excel file.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet.
            row_idx: 0-based row index to remove.

        Returns:
            bool: True if successful, False otherwise.
        """
        try:
            logger = get_logger()
            logger.info(f"Removing row {row_idx} from {file_path}, sheet {sheet_name}")

            # Create a backup
            backup_file = file_path + ".bak"
            copy2(file_path, backup_file)
            logger.debug(f"Created backup at {backup_file}")

            try:
                # Load workbook
                wb = load_workbook(file_path)
                ws = wb[sheet_name]

                # Convert 0-based row index to 1-based Excel row (add 2 for header)
                excel_row = row_idx + 2

                # Check if row exists
                if excel_row > ws.max_row:
                    logger.warning(f"Row {excel_row} does not exist (max row: {ws.max_row})")
                    return False

                # Delete the row
                ws.delete_rows(excel_row)
                logger.debug(f"Deleted Excel row {excel_row}")

                # Update table references if they exist
                for table in ws.tables.values():
                    try:
                        current_ref = table.ref
                        ref_parts = current_ref.split(":")
                        if len(ref_parts) == 2:
                            start_ref = ref_parts[0]
                            end_ref = ref_parts[1]

                            # Extract row numbers
                            end_row = int("".join(filter(str.isdigit, end_ref)))

                            # If the deleted row was within the table, adjust the table range
                            if excel_row <= end_row:
                                # Get column letters
                                start_col = "".join(filter(str.isalpha, start_ref))
                                end_col = "".join(filter(str.isalpha, end_ref))

                                # Update table reference
                                new_end_row = max(end_row - 1, int("".join(filter(str.isdigit, start_ref))))
                                table.ref = f"{start_col}1:{end_col}{new_end_row}"
                                logger.debug(f"Updated table reference to: {table.ref}")
                    except Exception as e:
                        logger.warning(f"Error updating table reference: {str(e)}")
                        continue

                # Save workbook
                wb.save(file_path)

                # Update our cached DataFrame if it exists
                if self.excel_data is not None and row_idx < len(self.excel_data):
                    self.excel_data = self.excel_data.drop(self.excel_data.index[row_idx]).reset_index(drop=True)
                    logger.debug(f"Updated cached DataFrame, new length: {len(self.excel_data)}")

                # Update hyperlink cache - remove entries for this row and shift others
                self._cache_lock.lockForWrite()
                try:
                    # Remove entries for the deleted row
                    keys_to_remove = [key for key in self._hyperlink_cache.keys() if isinstance(key, tuple) and key[0] == row_idx]
                    for key in keys_to_remove:
                        del self._hyperlink_cache[key]

                    # Shift entries for rows after the deleted row
                    keys_to_update = [(key, value) for key, value in self._hyperlink_cache.items()
                                    if isinstance(key, tuple) and key[0] > row_idx]
                    for (old_row, col), value in keys_to_update:
                        del self._hyperlink_cache[(old_row, col)]
                        self._hyperlink_cache[(old_row - 1, col)] = value

                    logger.debug("Updated hyperlink cache after row removal")
                finally:
                    self._cache_lock.unlock()

                # Remove backup after successful operation
                if path.exists(backup_file):
                    remove(backup_file)
                    logger.debug("Removed backup file after successful row removal")

                logger.info(f"Successfully removed row {row_idx} (Excel row {excel_row})")
                return True

            finally:
                if "wb" in locals():
                    wb.close()

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error removing row: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")

            # Try to restore from backup
            if "backup_file" in locals() and path.exists(backup_file):
                try:
                    copy2(backup_file, file_path)
                    logger.info("Restored from backup after error")
                except Exception as backup_e:
                    logger.error(f"Failed to restore from backup: {str(backup_e)}")
                    if hasattr(backup_e, "__traceback__"):
                        logger.error(f"Backup restore traceback: {traceback.format_exception(type(backup_e), backup_e, backup_e.__traceback__)}")

            return False
