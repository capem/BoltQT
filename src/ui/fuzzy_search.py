from __future__ import annotations

import os
import platform
import subprocess
import traceback
import urllib.parse
from difflib import SequenceMatcher
from typing import Any, Callable, List, Optional

from openpyxl import load_workbook
from PyQt6.QtCore import Qt, QTimer, pyqtSignal
from PyQt6.QtWidgets import (
    QLineEdit,
    QListWidget,
    QMenu,
    QVBoxLayout,
    QWidget,
)

from ..utils.logger import get_logger


class FuzzySearchFrame(QWidget):
    """Frame containing a fuzzy search entry and listbox."""

    # Signals
    value_selected = pyqtSignal()  # Emitted when a value is selected
    text_changed = pyqtSignal()  # Emitted when the text input changes

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        values: Optional[List[str]] = None,
        search_threshold: int = 65,
        identifier: Optional[str] = None,
        on_tab: Optional[Callable[[Any], Optional[str]]] = None,
    ) -> None:
        super().__init__(parent)

        self.all_values = [str(v) for v in (values or []) if v is not None]
        self.search_threshold = max(
            0, min(100, search_threshold)
        )  # Clamp between 0 and 100
        self.identifier = identifier or "unnamed"
        self.on_tab_callback = on_tab

        # Setup UI
        self._create_widgets()
        self._setup_styles()
        self._bind_events()
        self._update_listbox()

        # Update timer for debouncing
        self._update_timer = QTimer()
        self._update_timer.setSingleShot(True)
        self._update_timer.timeout.connect(self._update_listbox)

    def _create_widgets(self) -> None:
        """Create and configure all child widgets."""
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)  # Remove margins completely
        layout.setSpacing(0)  # Minimal spacing between widgets
        self.setLayout(layout)

        # Entry widget with placeholder
        self.entry = QLineEdit()
        self.entry.setPlaceholderText("Type to search...")
        # self.entry.setFixedHeight(24)  # Reduce height of entry field
        layout.addWidget(self.entry)

        # Listbox
        self.listbox = QListWidget()
        # Allow horizontal scrollbar when needed, but don't show it always
        self.listbox.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.listbox.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        # Set fixed height to prevent excessive vertical growth and resizing
        self.listbox.setFixedHeight(200)  # Set a fixed height for the listbox
        layout.addWidget(self.listbox)

    def _setup_styles(self) -> None:
        """Configure styles for the widgets."""
        self.entry.setStyleSheet("""
            QLineEdit {
                padding: 3px 6px;
                border: 1px solid #d1d1d1;
                border-radius: 1px;
                background-color: white;
                font-family: system-ui;
                font-size: 10pt;
            }
            QLineEdit:focus {
                border: 1px solid #007aff;
            }
        """)

        self.listbox.setStyleSheet("""
            QListWidget {
                border: 1px solid #e0e0e0;
                border-radius: 1px;
                background-color: white;
                font-family: system-ui;
                font-size: 10pt;
                outline: none;
            }
            QListWidget::item {
                padding: 1px 6px;
                border-bottom: none;
            }
            QListWidget::item:selected {
                background-color: #e6f2ff;
                color: #000000;
                border-left: 2px solid #007aff;
            }
            QListWidget::item:hover:!selected {
                background-color: #f5f5f5;
            }
        """)

    def _bind_events(self) -> None:
        """Bind event handlers to widgets."""
        # Entry events
        self.entry.textChanged.connect(self._on_text_changed)
        self.entry.textChanged.connect(
            lambda: self.text_changed.emit()
        )  # Emit text_changed signal
        self.entry.returnPressed.connect(self._select_top_match)

        # Listbox events
        self.listbox.itemClicked.connect(self._on_select)
        self.listbox.itemDoubleClicked.connect(self._on_select)
        self.listbox.customContextMenuRequested.connect(self._show_context_menu)

        # Make listbox properly handle tab/enter keys
        self.listbox.installEventFilter(self)

        # Enable context menu for listbox
        self.listbox.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)

    def _on_text_changed(self) -> None:
        """Handle text changes in the entry widget."""
        # Restart the update timer
        self._update_timer.start(100)  # 100ms debounce

    def _update_listbox(self) -> None:
        """Update the listbox with intelligent fuzzy search results."""
        current_value = self.get()

        # Clear current listbox
        self.listbox.clear()

        # If empty, show all values
        if not current_value:
            self.listbox.addItems(self.all_values)
            return

        try:
            search_lower = current_value.lower()
            scored_matches: List[tuple[float, str]] = []

            for value in self.all_values:
                value_lower = value.lower()

                # Calculate ratio using SequenceMatcher
                ratio = SequenceMatcher(None, search_lower, value_lower).ratio() * 100

                # Apply bonuses for special matches
                if value_lower == search_lower:  # Exact match
                    ratio = 100
                elif value_lower.startswith(search_lower):  # Prefix match
                    ratio = max(ratio, 90)
                elif search_lower in value_lower:  # Contains match
                    ratio = max(ratio, 80)
                elif any(
                    word.startswith(search_lower) for word in value_lower.split()
                ):  # Word boundary match
                    ratio = max(ratio, 75)

                # Only include matches that meet the threshold
                if ratio >= self.search_threshold:
                    scored_matches.append((ratio, value))

            # Sort by score (highest first) and add to listbox
            scored_matches.sort(reverse=True, key=lambda x: x[0])
            for _, value in scored_matches:
                self.listbox.addItem(value)

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error in fuzzy search ({self.identifier}): {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            # Fall back to simple contains matching
            for value in self.all_values:
                if current_value.lower() in value.lower():
                    self.listbox.addItem(value)

    def _on_select(self) -> None:
        """Handle selection in the listbox."""
        current_item = self.listbox.currentItem()
        if current_item:
            value = current_item.text()
            self.set(value)
            self.value_selected.emit()

    def _select_top_match(self) -> None:
        """Select the top match when Enter is pressed."""
        if self.listbox.count() > 0:
            value = self.listbox.item(0).text()
            self.set(value)
            self.value_selected.emit()

    def get(self) -> str:
        """Get the current entry text."""
        return self.entry.text()

    def set(self, value: str) -> None:
        """Set the entry text."""
        self.entry.setText(str(value))

    def set_values(self, values: Optional[List[str]]) -> None:
        """Update the list of searchable values."""
        self.all_values = [str(v) for v in (values or []) if v is not None]
        current_value = self.get()
        self.set(current_value)
        self._update_listbox()

    def clear(self) -> None:
        """Clear the entry text and listbox."""
        self.entry.clear()
        self.listbox.clear()

    def keyPressEvent(self, event) -> None:
        """Handle key press events."""
        if event.key() == Qt.Key.Key_Down and self.listbox.count() > 0:
            # Move focus to listbox and select first item
            self.listbox.setFocus()
            self.listbox.setCurrentRow(0)
        elif event.key() == Qt.Key.Key_Enter or event.key() == Qt.Key.Key_Return:
            # Select the top match but stay in this filter
            self._select_top_match()
            # Keep focus on this input
            self.entry.setFocus()
            event.accept()
            return
        elif event.key() == Qt.Key.Key_Tab and self.listbox.count() > 0:
            # Select the top match first
            self._select_top_match()
            # If tab handler is provided, call it to move focus
            if self.on_tab_callback:
                if self.on_tab_callback(event) == "break":
                    event.accept()
                    return

        super().keyPressEvent(event)

    def _show_context_menu(self, pos) -> None:
        """Show the context menu on right-click."""
        item = self.listbox.itemAt(pos)
        if not item:
            return

        value = item.text()
        if not value.startswith("✓ "):  # Only show menu for hyperlinked items
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #d1d1d1;
                border-radius: 4px;
                padding: 4px 0px;
            }
            QMenu::item {
                padding: 4px 24px;
                font-family: system-ui;
                font-size: 10pt;
            }
            QMenu::item:selected {
                background-color: #e6f2ff;
                color: #000000;
            }
        """)
        menu.addAction("Open Linked File", lambda: self._open_linked_file(value))
        menu.exec(self.listbox.mapToGlobal(pos))

    def _open_linked_file(self, value: str) -> None:
        """Open the linked file from the Excel hyperlink.

        Args:
            value: The formatted value containing the checkmark and row information
        """
        # Find the parent ProcessingTab
        processing_tab = self._get_processing_tab()
        if not processing_tab:
            logger = get_logger()
            logger.warning("Could not find ProcessingTab parent")
            return

        try:
            # Parse the value to get the row index
            _, row_idx = processing_tab._parse_filter2_value(value)
            if row_idx < 0:
                logger = get_logger()
                logger.warning(f"Invalid row index: {row_idx}")
                return

            # Get the configuration
            config = processing_tab.config_manager.get_config()
            excel_file = config.get("excel_file")
            excel_sheet = config.get("excel_sheet")
            filter2_column = config.get("filter2_column")

            if not all([excel_file, excel_sheet, filter2_column]):
                logger = get_logger()
                logger.warning("Missing required configuration")
                return

            # Get the hyperlink from Excel
            wb = load_workbook(excel_file, data_only=True)
            ws = wb[excel_sheet]

            # Find the column index
            header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
            if filter2_column not in header:
                logger = get_logger()
                logger.warning(f"Column '{filter2_column}' not found")
                return
            col_idx = header[filter2_column]

            # Get the cell
            cell = ws.cell(
                row=row_idx + 2, column=col_idx
            )  # +2 for header and 1-based index

            if not cell.hyperlink:
                logger = get_logger()
                logger.warning(f"No hyperlink found in cell {cell.coordinate}")
                return

            # Get the hyperlink target
            target = cell.hyperlink.target
            if not target:
                logger = get_logger()
                logger.warning("Empty hyperlink target")
                return

            # Resolve the path if it's relative
            # Fix 1: URL decode the target path to handle %20 and other encoded characters
            target = urllib.parse.unquote(target)

            if not os.path.isabs(target):
                target = os.path.join(os.path.dirname(excel_file), target)

            # Fix 2: Normalize path separators based on the platform
            # For Windows (when running on WSL or native Windows)
            if target.startswith("//") or target.startswith("\\\\"):
                # This is a UNC path - ensure consistent formatting
                logger = get_logger()
                logger.debug(f"Normalizing UNC path: {target}")

                if os.name == "nt":  # Native Windows
                    # Convert to Windows backslash format
                    target = target.replace("/", "\\")
                    # Ensure UNC prefix is correct
                    if not target.startswith("\\\\"):
                        target = "\\\\" + target[2:]
                    logger.debug(f"Normalized to Windows format: {target}")
                else:  # WSL or Linux
                    # Convert to forward slash format for WSL
                    target = target.replace("\\", "/")
                    # Ensure UNC prefix is correct
                    if not target.startswith("//"):
                        target = "//" + target[2:]
                    logger.debug(f"Normalized to Unix format: {target}")

            logger.debug(f"Final target path: {target}")

            # Open the file using the system's default application

            if platform.system() == "Windows":
                os.startfile(target)
            elif platform.system() == "Darwin":  # macOS
                subprocess.call(("open", target))
            else:  # Linux
                subprocess.call(("xdg-open", target))

            logger.info(f"Opened linked file: {target}")

        except Exception as e:
            logger = get_logger()
            logger.error(f"Error opening linked file: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")

    def _get_processing_tab(self):
        """Get the parent ProcessingTab instance."""
        parent = self.parent()
        while parent:
            if hasattr(parent, "_parse_filter2_value") and hasattr(
                parent, "config_manager"
            ):
                return parent
            parent = parent.parent()
        return None

    def eventFilter(self, obj, event) -> bool:
        """Filter events for the listbox."""
        if obj == self.listbox and event.type() == event.Type.KeyPress:
            if event.key() == Qt.Key.Key_Tab:
                # Select the current item
                current_item = self.listbox.currentItem()
                if current_item:
                    value = current_item.text()
                    self.set(value)
                    self.value_selected.emit()

                # If tab handler is provided, call it to move focus
                if self.on_tab_callback:
                    if self.on_tab_callback(event) == "break":
                        return True

            elif event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
                # Select the current item
                current_item = self.listbox.currentItem()
                if current_item:
                    value = current_item.text()
                    self.set(value)
                    self.value_selected.emit()
                    # Return focus to the entry
                    self.entry.setFocus()
                    return True

        return super().eventFilter(obj, event)
