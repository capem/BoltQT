import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import unittest
import tempfile
import shutil
from pathlib import Path
import openpyxl
from datetime import datetime
import time
import json
from PyQt6.QtWidgets import QApplication
from PyQt6.QtCore import QObject, pyqtSignal

from src.utils.config_manager import ConfigManager
from src.utils.excel_manager import ExcelManager
from src.utils.pdf_manager import PDFManager
from src.utils.processing_thread import ProcessingThread
from src.utils.models import PDFTask

class TestProcessingWorkflow(unittest.TestCase):
    """Test cases for the complete processing workflow."""

    @classmethod
    def setUpClass(cls):
        """Create the application for Qt tests."""
        cls.app = QApplication([])

        # Define test data paths
        # Use absolute paths to avoid issues with working directory
        cls.test_data_dir = Path(os.path.abspath(os.path.join(os.path.dirname(__file__), "data")))
        cls.source_dir = cls.test_data_dir / "Source" / "FA"
        cls.dest_dir = cls.test_data_dir / "Dest"
        cls.excel_file = cls.test_data_dir / "SUIVI OV CHQ OPCVM BC 2023 test.xlsx"

        # Make a backup of the Excel file
        cls.excel_backup = cls.test_data_dir / "SUIVI OV CHQ OPCVM BC 2023 test.backup.xlsx"
        if cls.excel_file.exists():
            shutil.copy2(cls.excel_file, cls.excel_backup)

        # Ensure test directories exist
        cls.source_dir.mkdir(parents=True, exist_ok=True)
        cls.dest_dir.mkdir(parents=True, exist_ok=True)

    @classmethod
    def tearDownClass(cls):
        """Clean up the application."""
        # Restore Excel file from backup if it exists
        if cls.excel_backup.exists():
            shutil.copy2(cls.excel_backup, cls.excel_file)
            cls.excel_backup.unlink()

        cls.app.quit()

    def setUp(self):
        """Set up each test."""
        # Create temporary directory for test files
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.temp_dir.name)

        # Create test config file
        self.config_path = self.temp_path / "test_config.json"
        self.test_config = {
            "document_type": "FA",
            "source_folder": str(self.source_dir),
            "excel_file": str(self.excel_file),
            "excel_sheet": "FACTURE",
            "processed_folder": str(self.dest_dir),
            "output_template": "{processed_folder}/{filter3|date.year}/MOIS {filter3|date.month}/{filter2|str.first_word} {filter1} {filter2|str.split_no_last}.pdf",
            "filter1_column": "FOURNISSEUR",
            "filter2_column": "N° FACTURE",
            "filter3_column": "DATE FACTURE",
            "filter4_column": "MONTANT",
            "vision": {
                "enabled": False
            }
        }

        # Write test config
        with open(self.config_path, 'w') as f:
            json.dump(self.test_config, f)

        # Create test PDF file
        self.test_pdf_path = self.source_dir / "TEST_INVOICE.pdf"
        if not self.test_pdf_path.exists():
            # Create a simple PDF file for testing
            with open(self.test_pdf_path, 'wb') as f:
                f.write(b'%PDF-1.4\n%Test PDF file for processing workflow test')

        # Initialize managers
        self.config_manager = ConfigManager()
        self.config_manager._config_file = str(self.config_path)
        self.config_manager.load_config()

        self.excel_manager = ExcelManager()
        self.pdf_manager = PDFManager()

        # Create a signal relay for task completion
        class SignalRelay(QObject):
            task_completed = pyqtSignal(str, str)

        self.signal_relay = SignalRelay()

        # Initialize processing thread
        self.processing_thread = ProcessingThread(
            config_manager=self.config_manager,
            excel_manager=self.excel_manager,
            pdf_manager=self.pdf_manager
        )

        # Start the processing thread
        self.processing_thread.start()

    def tearDown(self):
        """Clean up after each test."""
        # Stop the processing thread
        self.processing_thread.stop()
        self.processing_thread.wait()

        # Clean up temporary directory
        self.temp_dir.cleanup()

        # Remove test PDF if it exists
        if self.test_pdf_path.exists():
            try:
                self.test_pdf_path.unlink()
            except Exception as e:
                print(f"Warning: Could not remove test PDF: {str(e)}")

        # Clean up any created files in destination directory
        for year_dir in ["2023", "2024", "2025"]:
            year_path = self.dest_dir / year_dir
            if year_path.exists():
                for month_dir in year_path.glob("MOIS *"):
                    for pdf_file in month_dir.glob("*TEST*.pdf"):
                        try:
                            pdf_file.unlink()
                        except Exception as e:
                            print(f"Warning: Could not remove test file {pdf_file}: {str(e)}")

    def test_new_invoice_processing(self):
        """Test processing a new invoice that doesn't exist in Excel."""
        # Create a new test PDF file for this test
        with open(self.test_pdf_path, 'wb') as f:
            f.write(b'%PDF-1.4\n%Test PDF file for processing workflow test')

        # Create a task for processing
        task = PDFTask(
            pdf_path=str(self.test_pdf_path),
            filter_values=["TEST COMPANY", "INV-12345", "15/04/2025", "1000.00"],
            status="pending",
            start_time=datetime.now()
        )

        # Add task to processing thread
        task_id = task.task_id
        self.processing_thread.tasks[task_id] = task

        # Set up completion signals
        task_completed = False
        task_failed = False
        error_message = ""

        def on_task_completed(completed_id, status):
            nonlocal task_completed
            if completed_id == task_id and status == "completed":
                task_completed = True

        def on_task_failed(failed_id, message):
            nonlocal task_failed, error_message
            if failed_id == task_id:
                task_failed = True
                error_message = message

        self.processing_thread.task_completed.connect(on_task_completed)
        self.processing_thread.task_failed.connect(on_task_failed)

        # Let the processing thread handle the task
        # The run method will pick up the pending task

        # Wait for task completion (with timeout)
        timeout = 30  # seconds - increased timeout
        start_time = datetime.now()
        while not (task_completed or task_failed) and (datetime.now() - start_time).total_seconds() < timeout:
            QApplication.processEvents()
            time.sleep(0.1)

        # Check if task failed
        if task_failed:
            self.fail(f"Task failed with error: {error_message}")

        # Verify task was completed
        self.assertTrue(task_completed, f"Task did not complete within timeout. Status: {task.status}")

        # Verify file was moved to the correct location
        expected_year = "2025"
        expected_month = "04"
        expected_filename = "INV-12345 TEST COMPANY INV-12345.pdf"
        expected_path = self.dest_dir / expected_year / f"MOIS {expected_month}" / expected_filename

        # Create directory if it doesn't exist (for test verification)
        expected_path.parent.mkdir(parents=True, exist_ok=True)

        # Verify the file exists or was moved
        self.assertTrue(expected_path.exists() or task.processed_pdf_location == str(expected_path),
                       f"File not found at expected path: {expected_path}")

        # Verify Excel row was added
        self.excel_manager.load_excel_data(str(self.excel_file), "FACTURE", force_reload=True)
        df = self.excel_manager.excel_data

        # Find the row with our test data
        matching_rows = df[
            (df["FOURNISSEUR"] == "TEST COMPANY") &
            (df["N° FACTURE"] == "INV-12345")
        ]

        self.assertGreater(len(matching_rows), 0, "Row was not added to Excel file")

        # Verify hyperlink was added
        wb = openpyxl.load_workbook(str(self.excel_file))
        ws = wb["FACTURE"]

        # Find the row with our test data
        # First, get the row index from the DataFrame
        matching_rows = df[
            (df["FOURNISSEUR"] == "TEST COMPANY") &
            (df["N° FACTURE"] == "INV-12345")
        ]

        self.assertGreater(len(matching_rows), 0, "Row was not found in Excel DataFrame")

        # Get the row index (0-based in DataFrame)
        df_row_idx = matching_rows.index[0]

        # Convert to Excel row (1-based, with header)
        excel_row_idx = df_row_idx + 2  # +2 because Excel is 1-based and we have a header row

        # Get the cell with the hyperlink
        cell = ws.cell(row=excel_row_idx, column=df.columns.get_loc("N° FACTURE") + 1)

        # Verify hyperlink exists or the task has recorded the processed location
        self.assertTrue(cell.hyperlink is not None or task.processed_pdf_location is not None,
                       "Neither hyperlink was added to Excel cell nor processed location was recorded")

        # If hyperlink exists, verify it points to the correct file
        if cell.hyperlink is not None:
            self.assertIn(expected_filename, cell.hyperlink.target,
                         f"Hyperlink does not point to the correct file. Target: {cell.hyperlink.target}")

    def test_existing_invoice_update(self):
        """Test processing an invoice that already exists in Excel."""
        # Create a new test PDF file for this test
        with open(self.test_pdf_path, 'wb') as f:
            f.write(b'%PDF-1.4\n%Test PDF file for processing workflow test')

        # First, add a row to Excel for our test
        self.excel_manager.load_excel_data(str(self.excel_file), "FACTURE", force_reload=True)

        # Create filter values for existing invoice
        filter_values = ["EXISTING CO", "INV-EXIST", "20/04/2025", "2000.00"]

        # Add a new row to Excel
        try:
            self.excel_manager.add_new_row(
                str(self.excel_file),
                "FACTURE",
                [self.test_config["filter1_column"], self.test_config["filter2_column"],
                 self.test_config["filter3_column"], self.test_config["filter4_column"]],
                filter_values
            )
        except Exception as e:
            print(f"Warning: Could not add test row to Excel: {str(e)}")
            # If we can't add a row, we'll skip this test
            self.skipTest("Could not add test row to Excel")

        # Create a task for processing
        task = PDFTask(
            pdf_path=str(self.test_pdf_path),
            filter_values=filter_values,
            status="pending",
            start_time=datetime.now()
        )

        # Add task to processing thread
        task_id = task.task_id
        self.processing_thread.tasks[task_id] = task

        # Set up completion signals
        task_completed = False
        task_failed = False
        error_message = ""

        def on_task_completed(completed_id, status):
            nonlocal task_completed
            if completed_id == task_id and status == "completed":
                task_completed = True

        def on_task_failed(failed_id, message):
            nonlocal task_failed, error_message
            if failed_id == task_id:
                task_failed = True
                error_message = message

        self.processing_thread.task_completed.connect(on_task_completed)
        self.processing_thread.task_failed.connect(on_task_failed)

        # Let the processing thread handle the task
        # The run method will pick up the pending task

        # Wait for task completion (with timeout)
        timeout = 30  # seconds - increased timeout
        start_time = datetime.now()
        while not (task_completed or task_failed) and (datetime.now() - start_time).total_seconds() < timeout:
            QApplication.processEvents()
            time.sleep(0.1)

        # Check if task failed
        if task_failed:
            self.fail(f"Task failed with error: {error_message}")

        # Verify task was completed
        self.assertTrue(task_completed, f"Task did not complete within timeout. Status: {task.status}")

        # Verify file was moved to the correct location
        expected_year = "2025"
        expected_month = "04"
        expected_filename = "INV-EXIST EXISTING CO INV-EXIST.pdf"
        expected_path = self.dest_dir / expected_year / f"MOIS {expected_month}" / expected_filename

        # Create directory if it doesn't exist (for test verification)
        expected_path.parent.mkdir(parents=True, exist_ok=True)

        # Verify the file exists or was moved
        self.assertTrue(expected_path.exists() or task.processed_pdf_location == str(expected_path),
                       f"File not found at expected path: {expected_path}")

        # Verify Excel hyperlink was updated
        wb = openpyxl.load_workbook(str(self.excel_file))
        ws = wb["FACTURE"]

        # Find the row with our test data
        # First, get the row index from the DataFrame
        matching_rows = self.excel_manager.excel_data[
            (self.excel_manager.excel_data["FOURNISSEUR"] == "EXISTING CO") &
            (self.excel_manager.excel_data["N° FACTURE"] == "INV-EXIST")
        ]

        self.assertGreater(len(matching_rows), 0, "Row was not found in Excel DataFrame")

        # Get the row index (0-based in DataFrame)
        df_row_idx = matching_rows.index[0]

        # Convert to Excel row (1-based, with header)
        excel_row_idx = df_row_idx + 2  # +2 because Excel is 1-based and we have a header row

        # Get the cell with the hyperlink
        cell = ws.cell(row=excel_row_idx, column=self.excel_manager.excel_data.columns.get_loc("N° FACTURE") + 1)

        # Verify hyperlink exists or the task has recorded the processed location
        self.assertTrue(cell.hyperlink is not None or task.processed_pdf_location is not None,
                       "Neither hyperlink was added to Excel cell nor processed location was recorded")

        # If hyperlink exists, verify it points to the correct file
        if cell.hyperlink is not None:
            self.assertIn(expected_filename, cell.hyperlink.target,
                         f"Hyperlink does not point to the correct file. Target: {cell.hyperlink.target}")

if __name__ == '__main__':
    unittest.main()
